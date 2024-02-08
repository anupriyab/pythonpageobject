import openpyxl


def get_data(sheetName):
    workbook = openpyxl.load_workbook('..//Excel//TestData.xlsx')
    print(workbook.sheetnames)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    cols = sheet.max_column
    mainList = []

    for i in range(2, rows + 1):
        datalist = []
        for j in range(1, cols + 1):
            data = sheet.cell(row=i, column=j).value
            datalist.insert(j, data)
        mainList.insert(i, datalist)
    return mainList
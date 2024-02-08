import openpyxl

workbook = openpyxl.load_workbook('..//TestData.xlsx')
print(workbook.sheetnames)
sheet = workbook["Sheet1"]
row = sheet.max_row
col = sheet.max_column

for rows in range(1, row+1):
    for cols in range(1, col+1):
        print(sheet.cell(rows, cols).value,   end="   ")
    print()